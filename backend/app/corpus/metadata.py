from __future__ import annotations

import csv
from pathlib import Path
from typing import Any


FIELD_ALIASES: dict[str, str] = {
    "申请号": "application_number",
    "申请号/专利号": "application_number",
    "application_number": "application_number",
    "申请公布号": "publication_number",
    "公开号": "publication_number",
    "公开公告号": "publication_number",
    "publication_number": "publication_number",
    "授权公告号": "grant_number",
    "公告号": "grant_number",
    "grant_number": "grant_number",
    "专利名称": "title",
    "名称": "title",
    "题名": "title",
    "title": "title",
    "摘要": "abstract",
    "abstract": "abstract",
    "申请人": "applicants",
    "专利权人": "applicants",
    "applicant": "applicants",
    "applicants": "applicants",
    "发明人": "inventors",
    "inventor": "inventors",
    "inventors": "inventors",
    "IPC": "ipc",
    "IPC分类号": "ipc",
    "国际专利分类号": "ipc",
    "ipc": "ipc",
    "CPC": "cpc",
    "CPC分类号": "cpc",
    "cpc": "cpc",
    "申请日": "application_date",
    "application_date": "application_date",
    "公开日": "publication_date",
    "publication_date": "publication_date",
    "授权日": "grant_date",
    "公告日": "grant_date",
    "grant_date": "grant_date",
    "法律状态": "legal_status",
    "legal_status": "legal_status",
    "专利类型": "patent_type",
    "类型": "patent_type",
    "patent_type": "patent_type",
    "文件名": "source_file_name",
    "全文文件": "source_file_name",
    "file": "source_file_name",
    "filename": "source_file_name",
    "source_file_name": "source_file_name",
}

LIST_FIELDS = {"applicants", "inventors", "ipc", "cpc"}


def parse_metadata_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        return _parse_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported metadata table type: {suffix}")


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    raw = _read_table_text(path)
    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(raw.splitlines(), dialect=dialect)
    return [_normalize_row(row) for row in reader if any((value or "").strip() for value in row.values())]


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX metadata import requires openpyxl.") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    normalized: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        row = {header: "" if value is None else str(value).strip() for header, value in zip(headers, raw_row)}
        if any(row.values()):
            normalized.append(_normalize_row(row))
    return normalized


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key = FIELD_ALIASES.get(str(raw_key).strip(), str(raw_key).strip())
        value = "" if raw_value is None else str(raw_value).strip()
        if not value:
            continue
        normalized[key] = _split_list(value) if key in LIST_FIELDS else value
    if "source_file_name" in normalized:
        normalized["source_file_name"] = Path(str(normalized["source_file_name"])).name
    return normalized


def _split_list(value: str) -> list[str]:
    normalized = value.replace("；", ";").replace("，", ";").replace(",", ";").replace("、", ";").replace("|", ";")
    return [part.strip() for part in normalized.split(";") if part.strip()]


def _read_table_text(path: Path) -> str:
    for encoding in ["utf-8-sig", "utf-8", "gb18030"]:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("Metadata table encoding is not supported. Please save it as UTF-8 or GB18030.")
