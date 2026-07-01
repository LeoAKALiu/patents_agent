from __future__ import annotations

import csv
import hashlib
import re
import tempfile
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from uuid import uuid4

from openpyxl import load_workbook

from backend.app.knowledge.patent_search import (
    dedupe_patent_search_hits,
    normalize_publication_number,
)
from backend.app.knowledge.patent_sources import CNIPA_OFFICIAL_EXPORT_SOURCE
from backend.app.research.providers import sanitize_untrusted_text
from backend.app.schemas import CnipaExportImportFailure, CnipaExportImportResult, PatentSearchHit

FIELD_ALIASES: dict[str, str] = {
    "公开公告号": "publication_number",
    "申请公布号": "publication_number",
    "公开号": "publication_number",
    "授权公告号": "publication_number",
    "Publication Number": "publication_number",
    "publication_number": "publication_number",
    "申请号": "application_number",
    "Application Number": "application_number",
    "application_number": "application_number",
    "专利名称": "title",
    "名称": "title",
    "题名": "title",
    "发明名称": "title",
    "Title": "title",
    "title": "title",
    "申请人": "applicant",
    "专利权人": "applicant",
    "Applicant": "applicant",
    "applicant": "applicant",
    "发明人": "inventor",
    "Inventor": "inventor",
    "公开日": "publication_date",
    "公开公告日": "publication_date",
    "Publication Date": "publication_date",
    "publication_date": "publication_date",
    "申请日": "application_date",
    "Application Date": "application_date",
    "摘要": "abstract",
    "Abstract": "abstract",
    "abstract": "abstract",
    "IPC": "ipc",
    "国际分类号": "ipc",
    "CPC": "cpc",
    "权利要求": "claims",
    "Claims": "claims",
    "说明书": "description",
    "Description": "description",
    "链接": "url",
    "详情页": "url",
    "URL": "url",
    "url": "url",
}

LIST_FIELDS = {"ipc", "cpc"}
TABLE_SUFFIXES = {".csv", ".xlsx"}
ATTACHMENT_SUFFIXES = {".pdf", ".xml"}
MAX_ZIP_MEMBER_COUNT = 100
MAX_ZIP_MEMBER_BYTES = 10 * 1024 * 1024
MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES = 25 * 1024 * 1024
CN_PUBLICATION_NUMBER_PATTERN = re.compile(r"^CN\d{8,12}[A-Z]\d?$")
CN_APPLICATION_NUMBER_PATTERN = re.compile(r"^(?:CN)?(?:\d{8}|\d{12})(?:\.\d)?$")


@dataclass(frozen=True)
class CnipaExportImportContext:
    project_id: str
    plan_id: str
    import_ledger_id: str
    query: str
    strategy_group_id: str


def parse_cnipa_official_export_file(path: Path, *, context: CnipaExportImportContext) -> CnipaExportImportResult:
    raw_file_hash = _file_hash(path)
    if path.suffix.lower() == ".zip":
        return _parse_zip(path, context=context, raw_file_hash=raw_file_hash)
    rows = _parse_table(path)
    return _rows_to_result(
        rows,
        context=context,
        raw_file_hash=raw_file_hash,
        detected_schema=path.suffix.lower().lstrip("."),
        source_file_name=path.name,
    )


def _parse_zip(path: Path, *, context: CnipaExportImportContext, raw_file_hash: str) -> CnipaExportImportResult:
    warnings: list[str] = []
    failures: list[CnipaExportImportFailure] = []
    hits: list[PatentSearchHit] = []
    attachments: list[str] = []
    seen_attachments: set[str] = set()
    row_count = 0
    with tempfile.TemporaryDirectory() as tmpdir:
        with zipfile.ZipFile(path) as archive:
            members = [member for member in sorted(archive.infolist(), key=lambda item: item.filename) if not member.is_dir()]
            if len(members) > MAX_ZIP_MEMBER_COUNT:
                failures.append(
                    CnipaExportImportFailure(
                        source_file_name=path.name,
                        code="zip_member_limit_exceeded",
                        message=(
                            f"ZIP contains {len(members)} files; only the first {MAX_ZIP_MEMBER_COUNT} are inspected."
                        ),
                    )
                )
                members = members[:MAX_ZIP_MEMBER_COUNT]
            total_uncompressed_bytes = 0
            for member in members:
                if member.is_dir():
                    continue
                member_name = Path(member.filename).name
                suffix = Path(member_name).suffix.lower()
                if not member_name:
                    failures.append(
                        CnipaExportImportFailure(
                            source_file_name=path.name,
                            code="unsafe_zip_member_name",
                            message=f"ZIP member {member.filename!r} does not have a safe basename.",
                        )
                    )
                    continue
                if member.file_size > MAX_ZIP_MEMBER_BYTES:
                    failures.append(
                        CnipaExportImportFailure(
                            source_file_name=member_name,
                            code="zip_member_too_large",
                            message=(
                                f"{member_name} exceeds the {MAX_ZIP_MEMBER_BYTES} byte ZIP member limit and was skipped."
                            ),
                        )
                    )
                    if suffix in ATTACHMENT_SUFFIXES:
                        _append_attachment(attachments, seen_attachments, member_name)
                        warnings.append(
                            f"{member_name} 附件已识别但因大小限制被跳过；结果仅返回附件文件名，未生成候选。"
                        )
                    continue
                if total_uncompressed_bytes + member.file_size > MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES:
                    failures.append(
                        CnipaExportImportFailure(
                            source_file_name=member_name,
                            code="zip_total_size_limit_exceeded",
                            message=(
                                f"Skipping {member_name} because extracting it would exceed the "
                                f"{MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES} byte ZIP size limit."
                            ),
                        )
                    )
                    if suffix in ATTACHMENT_SUFFIXES:
                        _append_attachment(attachments, seen_attachments, member_name)
                        warnings.append(
                            f"{member_name} 附件已识别但因大小限制被跳过；结果仅返回附件文件名，未生成候选。"
                        )
                    continue
                total_uncompressed_bytes += member.file_size
                if suffix in TABLE_SUFFIXES:
                    child = Path(tmpdir) / f"{uuid4().hex}-{member_name}"
                    try:
                        child.write_bytes(_read_zip_member_bytes(archive, member))
                    except ValueError as exc:
                        failures.append(
                            CnipaExportImportFailure(
                                source_file_name=member_name,
                                code="zip_member_too_large",
                                message=str(exc),
                            )
                        )
                        continue
                    partial = _rows_to_result(
                        _parse_table(child),
                        context=context,
                        raw_file_hash=raw_file_hash,
                        detected_schema=f"zip:{suffix.lstrip('.')}",
                        source_file_name=member_name,
                    )
                    row_count += partial.row_count
                    hits.extend(partial.hits)
                    warnings.extend(partial.warnings)
                    failures.extend(partial.failures)
                elif suffix in ATTACHMENT_SUFFIXES:
                    _append_attachment(attachments, seen_attachments, member_name)
                    warnings.append(f"{member_name} 附件已识别；结果仅返回附件文件名，未生成候选。")
    deduped = dedupe_patent_search_hits(hits)
    return CnipaExportImportResult(
        import_ledger_id=context.import_ledger_id,
        raw_file_hash=raw_file_hash,
        detected_schema="zip",
        row_count=row_count,
        parsed_count=len(deduped),
        hits=deduped,
        attachments=attachments,
        warnings=warnings,
        failures=failures,
    )


def _parse_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _parse_csv(path)
    if suffix == ".xlsx":
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported CNIPA export table type: {suffix}")


def _parse_csv(path: Path) -> list[dict[str, Any]]:
    text = _read_text(path)
    try:
        dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.DictReader(text.splitlines(), dialect=dialect)
    return [_normalize_row(row) for row in reader if any((value or "").strip() for value in row.values())]


def _parse_xlsx(path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        row = {header: "" if value is None else str(value).strip() for header, value in zip(headers, raw_row)}
        if any(row.values()):
            parsed.append(_normalize_row(row))
    return parsed


def _rows_to_result(
    rows: list[dict[str, Any]],
    *,
    context: CnipaExportImportContext,
    raw_file_hash: str,
    detected_schema: str,
    source_file_name: str,
) -> CnipaExportImportResult:
    hits: list[PatentSearchHit] = []
    failures: list[CnipaExportImportFailure] = []
    for index, row in enumerate(rows, start=2):
        raw_publication_number = str(row.get("publication_number") or "")
        raw_application_number = str(row.get("application_number") or "")
        publication_number = _plausible_cn_publication_number(raw_publication_number)
        application_number = _plausible_cn_application_number(raw_application_number)
        title = sanitize_untrusted_text(str(row.get("title") or ""), max_len=300)
        abstract = sanitize_untrusted_text(str(row.get("abstract") or "")) or None
        if not (title or abstract):
            failures.append(
                CnipaExportImportFailure(
                    source_file_name=source_file_name,
                    row_number=index,
                    code="missing_required_fields",
                    message="CNIPA export row requires title or abstract.",
                )
            )
            continue
        if not (publication_number or application_number):
            code = "invalid_cn_identifier" if (raw_publication_number or raw_application_number) else "missing_required_fields"
            message = (
                "CNIPA export row requires a plausible CN publication/application identifier."
                if code == "invalid_cn_identifier"
                else "CNIPA export row requires publication/application number."
            )
            failures.append(
                CnipaExportImportFailure(
                    source_file_name=source_file_name,
                    row_number=index,
                    code=code,
                    message=message,
                )
            )
            continue
        metadata = {
            "raw_file_hash": raw_file_hash,
            "import_ledger_id": context.import_ledger_id,
            "source_file_name": source_file_name,
            "row_number": index,
            "strategy_group": context.strategy_group_id,
            "evidence_origin": "official_export",
        }
        for key in ["claims", "description", "inventor", "application_date"]:
            if row.get(key):
                metadata[key] = row[key]
        hits.append(
            PatentSearchHit(
                id=uuid4().hex,
                source=CNIPA_OFFICIAL_EXPORT_SOURCE,
                query=context.query,
                title=title or publication_number or application_number,
                url=str(row.get("url") or ""),
                publication_number=publication_number or None,
                application_number=application_number or None,
                applicant=str(row.get("applicant") or ""),
                publication_date=str(row.get("publication_date") or ""),
                abstract=abstract,
                ipc=list(row.get("ipc") or []),
                cpc=list(row.get("cpc") or []),
                metadata=metadata,
            )
        )
    deduped = dedupe_patent_search_hits(hits)
    return CnipaExportImportResult(
        import_ledger_id=context.import_ledger_id,
        raw_file_hash=raw_file_hash,
        detected_schema=detected_schema,
        row_count=len(rows),
        parsed_count=len(deduped),
        hits=deduped,
        failures=failures,
    )


def _plausible_cn_publication_number(value: str) -> str:
    normalized = normalize_publication_number(value)
    return normalized if CN_PUBLICATION_NUMBER_PATTERN.match(normalized) else ""


def _plausible_cn_application_number(value: str) -> str:
    normalized = normalize_publication_number(value)
    return normalized if CN_APPLICATION_NUMBER_PATTERN.match(normalized) else ""


def _append_attachment(attachments: list[str], seen_attachments: set[str], member_name: str) -> None:
    if member_name not in seen_attachments:
        attachments.append(member_name)
        seen_attachments.add(member_name)


def _read_zip_member_bytes(archive: zipfile.ZipFile, member: zipfile.ZipInfo) -> bytes:
    with archive.open(member) as handle:
        data = handle.read(MAX_ZIP_MEMBER_BYTES + 1)
    if len(data) > MAX_ZIP_MEMBER_BYTES:
        raise ValueError(f"{member.filename} exceeds the ZIP member size limit during extraction.")
    return data


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized: dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key = FIELD_ALIASES.get(str(raw_key).strip(), str(raw_key).strip())
        value = "" if raw_value is None else str(raw_value).strip()
        if not value:
            continue
        normalized[key] = _split_list(value) if key in LIST_FIELDS else value
    return normalized


def _split_list(value: str) -> list[str]:
    normalized = value.replace("；", ";").replace("，", ";").replace(",", ";").replace("、", ";").replace("|", ";")
    return [part.strip() for part in normalized.split(";") if part.strip()]


def _read_text(path: Path) -> str:
    for encoding in ["utf-8-sig", "utf-8", "gb18030"]:
        try:
            return path.read_text(encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise ValueError("CNIPA export table encoding is not supported. Please save it as UTF-8 or GB18030.")


def _file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()
