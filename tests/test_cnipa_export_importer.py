from pathlib import Path
from zipfile import ZipFile

import pytest
from openpyxl import Workbook

from backend.app.knowledge import cnipa_export
from backend.app.knowledge.cnipa_export import (
    CnipaExportImportContext,
    parse_cnipa_official_export_file,
)


def _context() -> CnipaExportImportContext:
    return CnipaExportImportContext(
        project_id="p-1",
        plan_id="plan-1",
        import_ledger_id="ledger-1",
        query="城市体检 智能体",
        strategy_group_id="cnipa-official-export",
    )


def test_parse_cnipa_csv_export_maps_fields(tmp_path: Path):
    path = tmp_path / "cnipa.csv"
    path.write_text(
        "公开公告号,专利名称,申请人,公开日,摘要,IPC\n"
        "CN112233445A,城市体检智能体任务编排方法,示例公司,2024-01-01,公开了一种任务编排方法。,G06Q\n",
        encoding="utf-8-sig",
    )

    result = parse_cnipa_official_export_file(path, context=_context())

    assert result.parsed_count == 1
    assert result.raw_file_hash
    assert result.hits[0].source == "cnipa_official_export"
    assert result.hits[0].publication_number == "CN112233445A"
    assert result.hits[0].title == "城市体检智能体任务编排方法"
    assert result.hits[0].applicant == "示例公司"
    assert result.hits[0].ipc == ["G06Q"]
    assert result.hits[0].metadata["raw_file_hash"] == result.raw_file_hash
    assert result.hits[0].metadata["import_ledger_id"] == "ledger-1"
    assert result.hits[0].metadata["source_file_name"] == "cnipa.csv"
    assert result.hits[0].metadata["row_number"] == 2


def test_parse_cnipa_xlsx_export_maps_fields(tmp_path: Path):
    path = tmp_path / "cnipa.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["申请公布号", "名称", "专利权人", "摘要"])
    sheet.append(["CN998877665A", "可信复核系统", "示例研究院", "公开了一种可信复核系统。"])
    workbook.save(path)

    result = parse_cnipa_official_export_file(path, context=_context())

    assert result.parsed_count == 1
    assert result.hits[0].publication_number == "CN998877665A"
    assert result.hits[0].title == "可信复核系统"
    assert result.hits[0].applicant == "示例研究院"


def test_parse_cnipa_zip_export_reads_nested_tables_and_warns_on_pdf(tmp_path: Path):
    csv_path = tmp_path / "inner.csv"
    csv_path.write_text(
        "申请号,题名,摘要\nCN202410000001,城市体检证据链方法,公开了一种证据链复核方法。\n",
        encoding="utf-8",
    )
    pdf_path = tmp_path / "scan.pdf"
    pdf_path.write_bytes(b"%PDF-1.4 scanned placeholder")
    zip_path = tmp_path / "cnipa.zip"
    with ZipFile(zip_path, "w") as archive:
        archive.write(csv_path, "metadata/inner.csv")
        archive.write(pdf_path, "docs/scan.pdf")

    result = parse_cnipa_official_export_file(zip_path, context=_context())

    assert result.row_count == 1
    assert result.parsed_count == 1
    assert result.hits[0].application_number == "CN202410000001"
    assert result.attachments == ["scan.pdf"]
    assert result.warnings == ["scan.pdf 附件已识别；结果仅返回附件文件名，未生成候选。"]


def test_parse_cnipa_export_rejects_rows_without_identifier_or_title(tmp_path: Path):
    path = tmp_path / "bad.csv"
    path.write_text("申请人,公开日\n示例公司,2024-01-01\n", encoding="utf-8")

    result = parse_cnipa_official_export_file(path, context=_context())

    assert result.parsed_count == 0
    assert result.failures[0].code == "missing_required_fields"


def test_parse_cnipa_export_rejects_non_cn_identifiers(tmp_path: Path):
    path = tmp_path / "bad-id.csv"
    path.write_text(
        "公开公告号,专利名称,摘要\nWO2024000123A1,城市体检智能体任务编排方法,公开了一种任务编排方法。\n",
        encoding="utf-8",
    )

    result = parse_cnipa_official_export_file(path, context=_context())

    assert result.parsed_count == 0
    assert result.failures[0].code == "invalid_cn_identifier"
    assert result.hits == []


def test_parse_cnipa_zip_export_skips_oversized_members_and_excess_member_count(tmp_path: Path, monkeypatch):
    monkeypatch.setattr(cnipa_export, "MAX_ZIP_MEMBER_COUNT", 2)
    monkeypatch.setattr(cnipa_export, "MAX_ZIP_MEMBER_BYTES", 160)
    monkeypatch.setattr(cnipa_export, "MAX_ZIP_TOTAL_UNCOMPRESSED_BYTES", 240)

    small_csv = tmp_path / "small.csv"
    small_csv.write_text(
        "申请号,题名,摘要\nCN202410000001,城市体检证据链方法,公开了一种证据链复核方法。\n",
        encoding="utf-8",
    )
    large_csv = tmp_path / "large.csv"
    large_csv.write_text(
        "申请号,题名,摘要\nCN202410000002,城市体检证据链方法,"
        + ("公开了一种证据链复核方法。" * 20)
        + "\n",
        encoding="utf-8",
    )
    ignored_csv = tmp_path / "ignored.csv"
    ignored_csv.write_text(
        "申请号,题名,摘要\nCN202410000003,不应处理的额外成员,公开了一种证据链复核方法。\n",
        encoding="utf-8",
    )
    zip_path = tmp_path / "unsafe.zip"
    with ZipFile(zip_path, "w") as archive:
        archive.write(large_csv, "a-large.csv")
        archive.write(small_csv, "b-small.csv")
        archive.write(ignored_csv, "c-ignored.csv")

    result = parse_cnipa_official_export_file(zip_path, context=_context())

    assert result.parsed_count == 1
    assert result.hits[0].application_number == "CN202410000001"
    assert {failure.code for failure in result.failures} == {
        "zip_member_limit_exceeded",
        "zip_member_too_large",
    }
    assert all(hit.application_number != "CN202410000003" for hit in result.hits)


def test_parse_cnipa_export_rejects_txt_and_xlsm_tables(tmp_path: Path):
    txt_path = tmp_path / "cnipa.txt"
    txt_path.write_text(
        "公开公告号,专利名称,摘要\nCN112233445A,城市体检任务编排方法,公开了一种任务编排方法。\n",
        encoding="utf-8",
    )
    xlsm_path = tmp_path / "cnipa.xlsm"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公开公告号", "专利名称", "摘要"])
    sheet.append(["CN112233446A", "城市体检证据链复核方法", "公开了一种证据链复核方法。"])
    workbook.save(xlsm_path)

    with pytest.raises(ValueError, match=r"Unsupported CNIPA export table type: \.txt"):
        parse_cnipa_official_export_file(txt_path, context=_context())
    with pytest.raises(ValueError, match=r"Unsupported CNIPA export table type: \.xlsm"):
        parse_cnipa_official_export_file(xlsm_path, context=_context())


def test_parse_cnipa_zip_export_rejects_txt_and_xlsm_members(tmp_path: Path):
    txt_path = tmp_path / "inner.txt"
    txt_path.write_text(
        "公开公告号,专利名称,摘要\nCN112233445A,城市体检任务编排方法,公开了一种任务编排方法。\n",
        encoding="utf-8",
    )
    xlsm_path = tmp_path / "inner.xlsm"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["公开公告号", "专利名称", "摘要"])
    sheet.append(["CN112233446A", "城市体检证据链复核方法", "公开了一种证据链复核方法。"])
    workbook.save(xlsm_path)
    zip_path = tmp_path / "cnipa.zip"
    with ZipFile(zip_path, "w") as archive:
        archive.write(txt_path, "metadata/inner.txt")
        archive.write(xlsm_path, "metadata/inner.xlsm")

    result = parse_cnipa_official_export_file(zip_path, context=_context())

    assert result.parsed_count == 0
    assert result.row_count == 0
    assert result.failures == []
